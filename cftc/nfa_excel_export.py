"""Export the currently filtered NFA firm list to Excel -- mirrors
excel_export.py's SEC-side pattern (dated filename, firms + principals on
separate sheets, auto-fit column widths)."""
from __future__ import annotations
from datetime import date

import pandas as pd
from openpyxl.utils import get_column_letter

from core import config
from cftc import nfa_db

EXPORTS_DIR = config.BASE_DIR / "exports" / "cftc"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def build_filename(states: list[str], reg_types: list[str]) -> str:
    from sec.excel_export import _slug  # reuse the same slugging rule

    parts = ["nfa_firms"]
    if states:
        parts.append(_slug("-".join(sorted(states))))
    else:
        parts.append("ALL")
    if reg_types:
        parts.append(_slug("-".join(sorted(reg_types))))
    parts.append(date.today().isoformat())
    return "_".join(parts) + ".xlsx"


def export_firms(df: pd.DataFrame, states: list[str], reg_types: list[str]) -> "Path":
    """Write the given (already-filtered) NFA firms DataFrame to a dated
    Excel file under exports/. Returns the file path."""
    export_cols = [
        "firm_name", "reg_types", "membership_status", "has_reg_actions",
        "city", "state", "country", "website", "website_source", "phone",
        "sec_prospect_id",
    ]
    cols = [c for c in export_cols if c in df.columns]
    out = df[cols].copy()
    if "sec_prospect_id" in out.columns:
        out["sec_prospect_id"] = out["sec_prospect_id"].notna()
    out = out.rename(columns={
        "firm_name": "Firm", "reg_types": "Registration Types",
        "membership_status": "Membership Status", "has_reg_actions": "Reg Actions",
        "city": "City", "state": "State", "country": "Country",
        "website": "Website", "website_source": "Website Source", "phone": "Phone",
        "sec_prospect_id": "Also SEC-Registered",
    })

    principal_cols = [
        "firm_name", "name", "title", "ten_percent_owner", "email",
        "email_verified", "email_source", "linkedin_profile_url",
    ]
    all_principals = pd.DataFrame([dict(r) for r in nfa_db.get_all_principals()])
    if not all_principals.empty:
        principals_out = all_principals[all_principals["firm_id"].isin(df["id"])][principal_cols].copy()
    else:
        principals_out = pd.DataFrame(columns=principal_cols)
    principals_out = principals_out.rename(columns={
        "firm_name": "Firm", "name": "Name", "title": "Title",
        "ten_percent_owner": "10%+ Owner", "email": "Email",
        "email_verified": "Verified", "email_source": "Email Source",
        "linkedin_profile_url": "Find This Person",
    })

    filename = build_filename(states, reg_types)
    path = EXPORTS_DIR / filename
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Firms")
        principals_out.to_excel(writer, index=False, sheet_name="Principals")

    from openpyxl import load_workbook
    wb = load_workbook(path)
    for sheet_name, sheet_df in (("Firms", out), ("Principals", principals_out)):
        ws = wb[sheet_name]
        for i, col in enumerate(sheet_df.columns, start=1):
            width = max(sheet_df[col].astype(str).map(len).max() if len(sheet_df) else 0, len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
    wb.save(path)

    return path


def export_full_database() -> "Path":
    """The whole NFA contact database, one row per PRINCIPAL (not per firm)
    -- mirrors sec/excel_export.py's export_full_database(). NFA has no AUM
    field (not a CFTC-disclosed concept the way SEC's Form ADV discloses it),
    so registration type + city/state stand in as the firm-level context.
    Ignores any dashboard filter -- always the full, current master list."""
    firms = [dict(r) for r in nfa_db.get_firms()]
    principals_by_firm: dict[int, list[dict]] = {}
    for p in nfa_db.get_all_principals():
        principals_by_firm.setdefault(p["firm_id"], []).append(dict(p))

    rows = []
    for f in firms:
        base = {
            "Firm": f["firm_name"], "Registration Types": f["reg_types"],
            "City": f["city"], "State": f["state"], "Website": f["website"],
            "Also SEC-Registered": bool(f["sec_prospect_id"]),
            "Reg Actions": f["has_reg_actions"], "Status": f.get("crm_stage"),
        }
        people = principals_by_firm.get(f["id"], [])
        if not people:
            rows.append({**base, "Name": None, "Title": None, "Email": None, "Verified": False, "10%+ Owner": False, "Find This Person": None})
        for p in people:
            rows.append({
                **base, "Name": p["name"], "Title": p["title"], "Email": p["email"],
                "Verified": bool(p["email_verified"]), "10%+ Owner": bool(p["ten_percent_owner"]),
                "Find This Person": p["linkedin_profile_url"],
            })

    cols = [
        "Firm", "Registration Types", "Name", "Title", "Email", "Verified",
        "10%+ Owner", "City", "State", "Website", "Also SEC-Registered",
        "Reg Actions", "Status", "Find This Person",
    ]
    out = pd.DataFrame(rows, columns=cols)

    filename = f"nfa_full_contact_database_{date.today().isoformat()}.xlsx"
    path = EXPORTS_DIR / filename
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="All Contacts")

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["All Contacts"]
    for i, col in enumerate(out.columns, start=1):
        width = max(out[col].astype(str).map(len).max() if len(out) else 0, len(col)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
    wb.save(path)

    return path
