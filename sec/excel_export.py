"""Export the currently filtered (and enriched) prospect list to Excel,
saved under exports/ with a naming convention that encodes the filter and date
— so repeated exports for different cuts don't collide or get confused."""
from __future__ import annotations
import re
from datetime import date

import pandas as pd
from openpyxl.utils import get_column_letter

from core import config
from sec import db
from core import formatting

EXPORTS_DIR = config.BASE_DIR / "exports" / "sec"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

_SLUG_RE = re.compile(r"[^a-zA-Z0-9]+")


def _slug(text: str) -> str:
    return _SLUG_RE.sub("-", text).strip("-")


def _format_aum(value: float) -> str:
    if value >= 1_000_000_000:
        return f"{value / 1_000_000_000:.1f}B"
    if value >= 1_000_000:
        return f"{value / 1_000_000:.0f}M"
    return f"{value:,.0f}"


def build_filename(states: list[str], cities: list[str], aum_range: tuple[float, float] | None) -> str:
    """e.g. prospects_NY_NewYork_AUM5M-500M_2026-07-11.xlsx"""
    parts = ["prospects"]

    if cities:
        parts.append(_slug("-".join(sorted(cities))[:40]))
    elif states:
        parts.append(_slug("-".join(sorted(states))))
    else:
        parts.append("ALL")

    if aum_range:
        parts.append(f"AUM{_format_aum(aum_range[0])}-{_format_aum(aum_range[1])}")

    parts.append(date.today().isoformat())
    return "_".join(parts) + ".xlsx"


def export_prospects(df: pd.DataFrame, states: list[str], cities: list[str], aum_range: tuple[float, float] | None) -> "Path":
    """Write the given (already-filtered) prospects DataFrame to a dated
    Excel file under exports/. Returns the file path."""
    export_cols = [
        "firm_name", "prospect_type", "hq_city", "hq_state", "aum",
        "contact_name", "contact_title", "email", "email_verified",
        "email_source", "website", "status", "notes",
    ]
    cols = [c for c in export_cols if c in df.columns]
    out = df[cols].copy()
    if "aum" in out.columns:
        out["aum"] = out["aum"].apply(formatting.format_aum)
    out = out.rename(columns={
        "firm_name": "Firm", "prospect_type": "Type", "hq_city": "City",
        "hq_state": "State", "aum": "AUM", "contact_name": "Contact",
        "contact_title": "Title", "email": "Email", "email_verified": "Verified",
        "email_source": "Email Source", "website": "Website", "status": "Status",
        "notes": "Notes",
    })

    # Secondary contacts (a firm can have more than one real contact, e.g.
    # several co-founders found on a team page) — kept on their own sheet
    # rather than exploding the main sheet into multiple rows per firm.
    contact_cols = [
        "firm_name", "contact_name", "contact_title", "email",
        "email_verified", "email_source", "linkedin_profile_url",
    ]
    all_contacts = pd.DataFrame([dict(r) for r in db.get_all_contacts()])
    if not all_contacts.empty:
        extra_out = all_contacts[all_contacts["prospect_id"].isin(df["id"])][contact_cols].copy()
    else:
        extra_out = pd.DataFrame(columns=contact_cols)
    extra_out = extra_out.rename(columns={
        "firm_name": "Firm", "contact_name": "Contact", "contact_title": "Title",
        "email": "Email", "email_verified": "Verified", "email_source": "Email Source",
        "linkedin_profile_url": "Find This Person",
    })

    filename = build_filename(states, cities, aum_range)
    path = EXPORTS_DIR / filename
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Prospects")
        extra_out.to_excel(writer, index=False, sheet_name="Additional Contacts")

    from openpyxl import load_workbook
    wb = load_workbook(path)
    for sheet_name, sheet_df in (("Prospects", out), ("Additional Contacts", extra_out)):
        ws = wb[sheet_name]
        for i, col in enumerate(sheet_df.columns, start=1):
            width = max(sheet_df[col].astype(str).map(len).max() if len(sheet_df) else 0, len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
    wb.save(path)

    return path


def _autofit(path, sheet_name: str, sheet_df: pd.DataFrame) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for i, col in enumerate(sheet_df.columns, start=1):
        width = max(sheet_df[col].astype(str).map(len).max() if len(sheet_df) else 0, len(col)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
    wb.save(path)


def export_full_database() -> "Path":
    """The whole SEC contact database, one row per CONTACT (not per firm) --
    every primary AND secondary contact gets its own row, each carrying the
    firm's AUM/location/website alongside it, so Mayank can filter/sort by
    AUM or state directly in Excel without cross-referencing a second sheet.
    Unlike export_prospects(), this ignores the dashboard's current filter --
    it's the full, always-current master list, not a scoped cut."""
    prospects = [dict(r) for r in db.get_all_prospects()]
    contacts_by_prospect: dict[int, list[dict]] = {}
    for c in db.get_all_contacts():
        contacts_by_prospect.setdefault(c["prospect_id"], []).append(dict(c))

    rows = []
    for p in prospects:
        base = {
            "Firm": p["firm_name"], "Type": p["prospect_type"], "AUM": formatting.format_aum(p["aum"]),
            "City": p["hq_city"], "State": p["hq_state"], "Website": p["website"],
            "CRD": p.get("crd_number"), "Status": p["status"],
        }
        if p["contact_name"] or p["email"]:
            rows.append({
                **base, "Contact": p["contact_name"], "Title": p["contact_title"],
                "Email": p["email"], "Verified": bool(p["email_verified"]),
                "Email Source": p["email_source"], "Is Primary Contact": True,
                "Find This Person": p["linkedin_profile_url"],
            })
        for c in contacts_by_prospect.get(p["id"], []):
            rows.append({
                **base, "Contact": c["contact_name"], "Title": c["contact_title"],
                "Email": c["email"], "Verified": bool(c["email_verified"]),
                "Email Source": c["email_source"], "Is Primary Contact": False,
                "Find This Person": c["linkedin_profile_url"],
            })
        if not p["contact_name"] and not p["email"] and not contacts_by_prospect.get(p["id"]):
            # No contact found at all for this firm -- still worth listing
            # (AUM/location alone is useful), just with blank contact fields.
            rows.append({
                **base, "Contact": None, "Title": None, "Email": None,
                "Verified": False, "Email Source": None, "Is Primary Contact": False,
                "Find This Person": None,
            })

    cols = [
        "Firm", "Type", "Contact", "Title", "Email", "Verified", "Email Source",
        "Is Primary Contact", "AUM", "City", "State", "Website", "CRD", "Status", "Find This Person",
    ]
    out = pd.DataFrame(rows, columns=cols)

    filename = f"sec_full_contact_database_{date.today().isoformat()}.xlsx"
    path = EXPORTS_DIR / filename
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="All Contacts")
    _autofit(path, "All Contacts", out)

    return path
